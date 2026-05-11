#!/usr/bin/env python3
"""Convert docs/requerimientos.md into docs/Requerimientos_actualizado.xlsx."""

from __future__ import annotations

import importlib.util
import re
import subprocess
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


ROOT_DIR = Path(__file__).resolve().parents[1]
VENV_PYTHON = Path("/tmp/xlsx-venv/bin/python3")
INPUT_MD = ROOT_DIR / "docs" / "requerimientos.md"
REFERENCE_XLSX = ROOT_DIR / "docs" / "Requerimientos.xlsx"
OUTPUT_XLSX = ROOT_DIR / "docs" / "Requerimientos_actualizado.xlsx"
EXPECTED_MODULES = 27
FIXED_XLSX_TIMESTAMP = (2026, 1, 1, 0, 0, 0)

MODULE_TO_SHEET = {
    "MÓDULO LANDING": "Landing",
    "Login": "Login",
    "MÓDULO RECUPERAR PASS": "Recuperar Pass",
    "MÓDULO REGISTRO": "Registro",
    "MÓDULO REGISTRO VENDEDOR": "Formulario Productor",
    "MÓDULO CIUDADES / UBICACIÓN": "Ciudades / Ubicación",
    "MÓDULO CATÁLOGO": "Catalogo",
    "MÓDULO DASHBOARD MIS PRODUCTOS": "Dashboard Productor",
    "MÓDULO PERFIL": "Dashboard Mi Perfil",
    "MÓDULO RESÉNAS": "Reseñas",
    "MÓDULO PAGINA 404": "404",
    "MÓDULO ENLACES EXTERNOS": "Enlaces Externos",
    "MÓDULO PRODUCTO DETALLE": "Producto detalle",
    "MÓDULO CARRITO": "Carrito",
    "MÓDULO CHECKOUT": "Checkout",
    "MÓDULO PEDIDOS": "Pedidos",
    "MÓDULO FAVORITOS": "Favoritos",
    "MÓDULO STAND DETALLE": "Stand detalle",
    "MÓDULO STANDS": "Stand directorio",
    "BASE DE DATOS": "Generales",
    "MÓDULO POLÍTICA PRIVACIDAD": "Política Privacidad",
    "MÓDULO TÉRMINOS Y CONDICIONES": "Términos y Condiciones",
    "MÓDULO NOTIFICACIONES": "Notificaciones",
    "MÓDULO STAND (MIS PRODUCTOS)": "Stand (Mis Productos)",
    "MÓDULO ADMIN DASHBOARD": "Admin Dashboard",
    "MÓDULO WORKERS": "Workers",
    "MÓDULO SERVICES & HELPERS": "Services & Helpers",
}

COLUMN_WIDTHS = {"A": 5, "B": 8, "C": 13, "D": 153, "E": 5, "F": 8, "G": 8, "H": 8, "I": 20}
LEGEND_ROWS = [
    ("SI", "Marque con una equis (X) si el sistema cumple con la función descrita"),
    ("NO", "Marque con una equis (X) si el sistema no cumple con la función descrita"),
    ("ADI", "Marque con una equis (X) si el sistema debe adicionar la función descrita"),
    ("Observaciones:", "Espacio para complementar respuestas"),
]
DATA_HEADERS = ["#", "Descripción", "SI/NO/ADI", "Nivel de Cumplimiento", "Observaciones"]
ILLEGAL_SHEET_CHARS = re.compile(r"[\\*?:\[\]]")
warnings: list[str] = []


def warn(message: str) -> None:
    warnings.append(message)
    print(f"Warning: {message}", file=sys.stderr)


def fail(message: str) -> None:
    print(f"Error: {message}", file=sys.stderr)
    sys.exit(1)


def preflight() -> None:
    if not VENV_PYTHON.exists():
        fail(
            f"Missing venv Python: {VENV_PYTHON}. Create it with: "
            "python3 -m venv /tmp/xlsx-venv && /tmp/xlsx-venv/bin/pip install openpyxl"
        )
    if not INPUT_MD.exists():
        fail(f"Missing input file: {INPUT_MD}")
    if not REFERENCE_XLSX.exists():
        fail(f"Missing format reference file: {REFERENCE_XLSX}")

    probe = subprocess.run(
        [str(VENV_PYTHON), "-c", "import openpyxl"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    if probe.returncode != 0:
        fail(f"openpyxl is not installed in {VENV_PYTHON}. Install with: /tmp/xlsx-venv/bin/pip install openpyxl")
    if importlib.util.find_spec("openpyxl") is None:
        fail(f"Current Python cannot import openpyxl. Run with: {VENV_PYTHON} {Path(__file__).as_posix()}")


def split_markdown_row(line: str) -> list[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]

    cells: list[str] = []
    current: list[str] = []
    escaped = False
    for char in text:
        if char == "|" and not escaped:
            cells.append("".join(current).strip().replace(r"\|", "|"))
            current = []
            escaped = False
            continue
        current.append(char)
        escaped = char == "\\" and not escaped
        if char != "\\":
            escaped = False
    cells.append("".join(current).strip().replace(r"\|", "|"))
    return cells


def is_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in cells)


def is_header_row(cells: list[str]) -> bool:
    normalized = [cell.strip().lower() for cell in cells]
    return normalized[:5] == ["#", "descripción", "si/no/adi", "nivel de cumplimiento", "observaciones"]


def normalize_requirement_cells(header: str, line: str, cells: list[str]) -> list[str] | None:
    if len(cells) == 5:
        return cells
    if len(cells) > 5:
        normalized = [cells[0], " | ".join(cells[1:-3]), cells[-3], cells[-2], cells[-1]]
        warn(f"{header}: normalized row with unescaped pipe characters: {line[:120]}")
        return normalized
    warn(f"{header}: skipped malformed row with {len(cells)} columns: {line[:120]}")
    return None


def detect_nuevo(row: dict[str, str]) -> bool:
    return "[NUEVO]" in row["observaciones"]


def is_section_header(number: str, description: str) -> bool:
    """Detect section header rows like REQUERIMIENTOS FUNCIONALES / NO FUNCIONALES."""
    return (not number.strip()
            and ("REQUERIMIENTOS" in description.upper()
                 or "REQUERIMIENTO" in description.upper()))


def parse_requirements(header: str, section: str) -> list[dict[str, Any]]:
    requirements: list[dict[str, Any]] = []
    in_requirements = False

    for raw_line in section.splitlines():
        line = raw_line.strip()
        if line == "### Requerimientos":
            in_requirements = True
            continue
        if in_requirements and (line.startswith("### ") or line == "---"):
            break
        if not in_requirements or not line.startswith("|"):
            continue

        cells = split_markdown_row(line)
        if is_separator_row(cells) or is_header_row(cells):
            continue
        normalized_cells = normalize_requirement_cells(header, line, cells)
        if normalized_cells is None:
            continue

        number, description, compliance, level, observations = normalized_cells
        if not any(cell.strip() for cell in cells):
            continue
        if not number.strip() and not description.strip() and not compliance.strip() and (level.strip() or observations.strip()):
            continue

        row = {
            "number": number.strip(),
            "description": description.strip(),
            "compliance": compliance.strip(),
            "level": level.strip(),
            "observaciones": observations.strip(),
        }
        row["is_nuevo"] = detect_nuevo(row)
        row["is_section_header"] = is_section_header(number, description)
        requirements.append(row)

    if not in_requirements:
        warn(f"{header}: missing ### Requerimientos section; sheet will be created without data rows")
    return requirements


def parse_modules(md_text: str) -> list[dict[str, Any]]:
    matches = list(re.finditer(r"^##\s+(.+?)\s*$", md_text, flags=re.MULTILINE))
    modules: list[dict[str, Any]] = []
    for index, match in enumerate(matches):
        header = match.group(1).strip()
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(md_text)
        section = md_text[start:end]
        if header not in MODULE_TO_SHEET:
            warn(f"Unmapped module header skipped: {header}")
            continue
        modules.append({"header": header, "sheet_name": MODULE_TO_SHEET[header], "requirements": parse_requirements(header, section)})

    if len(modules) != EXPECTED_MODULES:
        fail(f"Expected {EXPECTED_MODULES} mapped modules, parsed {len(modules)}")
    return modules


def sanitized_sheet_name(name: str, used: set[str]) -> str:
    sanitized = ILLEGAL_SHEET_CHARS.sub("", name).replace("/", " ").strip() or "Sheet"
    if sanitized != name:
        warn(f"Sanitized sheet name '{name}' to '{sanitized}'")
    if len(sanitized) > 31:
        warn(f"Truncated sheet name '{sanitized}' to 31 characters")
        sanitized = sanitized[:31]

    base = sanitized
    suffix = 2
    while sanitized in used:
        suffix_text = f"_{suffix}"
        sanitized = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        warn(f"Duplicate sheet name '{base}' renamed to '{sanitized}'")
        suffix += 1
    used.add(sanitized)
    return sanitized


def create_sheet(wb: Any, module: dict[str, Any], used_names: set[str]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet(title=sanitized_sheet_name(module["sheet_name"], used_names))
    font_9 = Font(name="Calibri", size=9)
    font_11 = Font(name="Calibri", size=11)
    font_11_bold = Font(name="Calibri", size=11, bold=True)
    font_14_bold = Font(name="Calibri", size=14, bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    nuevo_fill = PatternFill(fill_type="solid", fgColor="FFCC80")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    # Row 2: Title merged B2:I2
    ws.merge_cells("B2:I2")
    ws["B2"].value = module["header"]
    ws["B2"].font = font_14_bold
    ws["B2"].alignment = center
    ws["B2"].fill = title_fill
    ws.row_dimensions[2].height = 25

    # Rows 3-6: Legend in B-C
    for row_number, (key, meaning) in enumerate(LEGEND_ROWS, start=3):
        for column, value in ((2, key), (3, meaning)):
            cell = ws.cell(row=row_number, column=column, value=value)
            cell.font = font_9
            cell.alignment = left

    # Find the first section header from parsed data to use in row 9
    first_section = "REQUERIMIENTOS FUNCIONALES"
    for req in module["requirements"]:
        if req["is_section_header"]:
            first_section = req["description"]
            break

    # Row 8: "Nivel de cumplimiento" merged F8:H8, "OBSERVACIONES" at I8
    ws.merge_cells("F8:H8")
    ws["F8"].value = "Nivel de cumplimiento"
    ws["F8"].font = font_11_bold
    ws["F8"].alignment = center
    ws["F8"].border = thin_border
    for col_letter in ("G", "H"):
        ws[col_letter + "8"].border = thin_border
    ws["I8"].value = "OBSERVACIONES"
    ws["I8"].font = font_11_bold
    ws["I8"].alignment = center
    ws["I8"].border = thin_border

    # Row 9: Section header (merged C:D) + SI/NO/ADI column headers
    ws.merge_cells("C9:D9")
    ws["C9"].value = first_section
    ws["C9"].font = font_11_bold
    ws["C9"].alignment = left
    ws["C9"].border = thin_border
    ws["D9"].border = thin_border
    for col_letter, val in (("F", "SI"), ("G", "NO"), ("H", "ADI")):
        cell = ws[col_letter + "9"]
        cell.value = val
        cell.font = font_11_bold
        cell.alignment = center
        cell.border = thin_border
    ws["I9"].value = "OBSERVACIONES"
    ws["I9"].font = font_11_bold
    ws["I9"].alignment = center
    ws["I9"].border = thin_border

    # Data rows start at row 10; skip first section header (already in row 9)
    row_index = 10
    skipped_first_section = False
    for requirement in module["requirements"]:
        number = requirement["number"]
        description = requirement["description"]
        level = requirement["level"]
        compliance = requirement["compliance"]
        observaciones = requirement["observaciones"]
        is_nuevo = requirement["is_nuevo"]
        is_section = requirement["is_section_header"]

        # Skip the very first section header (already displayed in row 9)
        if is_section and not skipped_first_section:
            skipped_first_section = True
            continue

        ws.row_dimensions[row_index].height = 20

        if is_section:
            # Section header row: merged C:D
            ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
            cell_c = ws.cell(row=row_index, column=3, value=description)
            cell_c.font = font_11_bold
            cell_c.alignment = left
            cell_c.border = thin_border
            ws.cell(row=row_index, column=4).border = thin_border
            for col in (1, 2, 5, 6, 7, 8, 9):
                ws.cell(row=row_index, column=col).border = thin_border
        else:
            # Regular data row
            # Column C (3): number
            c = ws.cell(row=row_index, column=3)
            try:
                c.value = int(number) if number.strip() else None
            except ValueError:
                c.value = number
            c.font = font_11
            c.alignment = center
            c.border = thin_border

            # Column D (4): description
            c = ws.cell(row=row_index, column=4, value=description)
            c.font = font_11
            c.alignment = left
            c.border = thin_border

            # Columns A, B, E: borders only
            for col in (1, 2, 5):
                c = ws.cell(row=row_index, column=col)
                c.border = thin_border

            # Column F (6): SI checkmark
            c = ws.cell(row=row_index, column=6)
            if level.strip() == "x" or compliance.upper() == "SI":
                c.value = "x"
            c.font = font_11
            c.alignment = center
            c.border = thin_border

            # Column G (7): NO checkmark
            c = ws.cell(row=row_index, column=7)
            if compliance.upper() == "NO":
                c.value = "x"
            c.font = font_11
            c.alignment = center
            c.border = thin_border

            # Column H (8): ADI checkmark
            c = ws.cell(row=row_index, column=8)
            if compliance.upper() == "ADI":
                c.value = "x"
            c.font = font_11
            c.alignment = center
            c.border = thin_border

            # Column I (9): Observaciones
            c = ws.cell(row=row_index, column=9, value=observaciones if observaciones else None)
            c.font = font_11
            c.alignment = left
            c.border = thin_border

            # NUEVO highlighting
            if is_nuevo:
                for col in range(1, 10):
                    ws.cell(row=row_index, column=col).fill = nuevo_fill

        row_index += 1

    ws.freeze_panes = "A10"


def build_workbook(modules: list[dict[str, Any]]) -> Any:
    from openpyxl import Workbook

    wb = Workbook()
    fixed_datetime = datetime(*FIXED_XLSX_TIMESTAMP, tzinfo=timezone.utc)
    wb.properties.created = fixed_datetime
    wb.properties.modified = fixed_datetime
    wb.properties.creator = "VIVA"
    wb.properties.lastModifiedBy = "VIVA"
    wb.remove(wb.active)
    used_names: set[str] = set()
    for module in modules:
        create_sheet(wb, module, used_names)
    return wb


def normalize_xlsx_zip(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as source:
        entries = []
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "docProps/core.xml":
                payload = re.sub(
                    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-01-01T00:00:00Z</dcterms:modified>',
                    payload,
                )
            entries.append((info.filename, payload))

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for filename, payload in sorted(entries):
            info = zipfile.ZipInfo(filename, FIXED_XLSX_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            target.writestr(info, payload)


def main() -> None:
    try:
        preflight()
        md_text = INPUT_MD.read_text(encoding="utf-8", errors="replace")
        modules = parse_modules(md_text)
        wb = build_workbook(modules)
        OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUTPUT_XLSX)
        normalize_xlsx_zip(OUTPUT_XLSX)
        total_requirements = sum(len(module["requirements"]) for module in modules)
        total_nuevos = sum(1 for module in modules for row in module["requirements"] if row["is_nuevo"])
        print(f"Created {OUTPUT_XLSX}")
        print(f"Sheets: {len(wb.sheetnames)}")
        print(f"Rows parsed: {total_requirements}; NUEVO rows: {total_nuevos}; warnings: {len(warnings)}")
    except SystemExit:
        raise
    except Exception as exc:
        fail(f"Failed to convert markdown to Excel: {exc}")


if __name__ == "__main__":
    main()
